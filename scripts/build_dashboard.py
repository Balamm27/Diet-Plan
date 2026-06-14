from __future__ import annotations

import json
import re
from datetime import datetime
from html import escape
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/bala/Documents/Codex/Diet Plan")
WORKBOOK_PATH = ROOT / "Diet Plan.xlsx"
SITE_PATH = ROOT / "index.html"
DATA_PATH = ROOT / "site-data.json"

DAY_NAMES = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]

PANTRY_BASICS = ["Salt", "Black Pepper", "Water", "Neutral Oil", "Olive Oil"]

CANONICAL_OVERRIDES = {
    "mangoes": "mango",
    "mango": "mango",
    "bananas": "banana",
    "banana": "banana",
    "coocnut water": "coconut water",
    "coconut water": "coconut water",
    "dragon fruit": "dragon fruit",
    "blue berries": "blueberries",
    "blueberries": "blueberries",
    "brocolli": "broccoli",
    "broccoli": "broccoli",
    "chicken": "chicken breast",
    "chicken breast": "chicken breast",
    "chick pea": "chickpeas",
    "chickpeas": "chickpeas",
    "walnut": "walnuts",
    "walnuts": "walnuts",
    "coriander": "coriander",
    "mint": "mint",
    "basil": "basil",
    "thyme": "thyme",
    "sesame seeds": "sesame seeds",
    "coconut oil": "coconut oil",
    "olive oil": "olive oil",
    "navel orange": "navel orange",
    "beans": "green beans",
    "cauliflower": "cauliflower",
    "lady's finger": "okra",
    "lady's finger okra": "okra",
    "okra": "okra",
    "celery juice": "celery",
    "flax seed water": "flax seeds",
    "sweet potatoes": "sweet potatoes",
}

DISPLAY_NAMES = {
    "mango": "Mango",
    "banana": "Banana",
    "coconut water": "Coconut Water",
    "dragon fruit": "Dragon Fruit",
    "blueberries": "Blueberries",
    "chicken breast": "Chicken Breast",
    "coconut oil": "Coconut Oil",
    "navel orange": "Navel Orange",
    "broccoli": "Broccoli",
    "olive oil": "Olive Oil",
    "sesame seeds": "Sesame Seeds",
    "sweet potatoes": "Sweet Potatoes",
    "chickpeas": "Chickpeas",
    "walnuts": "Walnuts",
    "coriander": "Coriander",
    "mint": "Mint",
    "basil": "Basil",
    "thyme": "Thyme",
    "green beans": "Green Beans",
    "cauliflower": "Cauliflower",
    "okra": "Lady's Finger (Okra)",
    "celery": "Celery",
    "flax seeds": "Flax Seeds",
}

DISH_VISUALS = {
    "flax-seed-water": {"icon": "💧", "avatar": "FS"},
    "celery-juice": {"icon": "🥬", "avatar": "CJ"},
    "mango-smoothie": {"icon": "🥭", "avatar": "MS"},
    "berry-smoothie": {"icon": "🫐", "avatar": "BS"},
    "orange-chicken": {"icon": "🍊", "avatar": "OC"},
    "sweet-potatoes": {"icon": "🍠", "avatar": "SP"},
    "beans": {"icon": "🫛", "avatar": "GB"},
    "cauliflower": {"icon": "🥦", "avatar": "CF"},
    "ladys-finger": {"icon": "✨", "avatar": "OK"},
}

RECIPE_AMOUNT_PLANS = {
    ("flax-seed-water", "Flax Seeds"): ["1 tbsp", "2 tbsp", "3 tbsp", "4 tbsp", "5 tbsp", "6 tbsp"],
    ("flax-seed-water", "Water"): ["12 to 14 oz", "24 to 28 oz", "36 to 42 oz", "48 to 56 oz", "60 to 70 oz", "72 to 84 oz"],
    ("celery-juice", "Celery"): ["4 to 5 stalks", "8 to 10 stalks", "12 to 15 stalks", "16 to 20 stalks", "20 to 25 stalks", "24 to 30 stalks"],
    ("celery-juice", "Water"): ["2 to 3 tbsp", "1/4 cup", "1/3 cup", "1/2 cup", "2/3 cup", "3/4 cup"],
    ("mango-smoothie", "Mango"): ["1 heaping cup", "2 heaping cups", "3 cups", "4 cups", "5 cups", "6 cups"],
    ("mango-smoothie", "Banana"): ["1 medium", "2 medium", "3 medium", "4 medium", "5 medium", "6 medium"],
    ("mango-smoothie", "Coconut Water"): ["3/4 to 1 cup", "1 1/2 to 2 cups", "2 1/4 to 3 cups", "3 to 4 cups", "3 3/4 to 5 cups", "4 1/2 to 6 cups"],
    ("berry-smoothie", "Dragon Fruit"): ["1 cup", "2 cups", "3 cups", "4 cups", "5 cups", "6 cups"],
    ("berry-smoothie", "Blueberries"): ["3/4 cup", "1 1/2 cups", "2 1/4 cups", "3 cups", "3 3/4 cups", "4 1/2 cups"],
    ("berry-smoothie", "Banana"): ["1 medium", "2 medium", "3 medium", "4 medium", "5 medium", "6 medium"],
    ("berry-smoothie", "Coconut Water"): ["3/4 cup", "1 1/2 cups", "2 1/4 cups", "3 cups", "3 3/4 cups", "4 1/2 cups"],
    ("orange-chicken", "Chicken Breast"): ["6 to 8 oz", "12 to 16 oz", "18 to 24 oz", "24 to 32 oz", "30 to 40 oz", "36 to 48 oz"],
    ("orange-chicken", "Navel Orange"): ["1 whole", "2 whole", "3 whole", "4 whole", "5 whole", "6 whole"],
    ("orange-chicken", "Broccoli"): ["1 1/2 to 2 cups florets", "3 to 4 cups florets", "4 1/2 to 6 cups florets", "6 to 8 cups florets", "7 1/2 to 10 cups florets", "9 to 12 cups florets"],
    ("orange-chicken", "Coconut Oil"): ["1 tsp", "2 tsp", "1 tbsp", "4 tsp", "5 tsp", "2 tbsp"],
    ("orange-chicken", "Olive Oil"): ["1 tsp", "2 tsp", "1 tbsp", "4 tsp", "5 tsp", "2 tbsp"],
    ("orange-chicken", "Sesame Seeds"): ["1 tsp", "2 tsp", "1 tbsp", "4 tsp", "5 tsp", "2 tbsp"],
    ("sweet-potatoes", "Sweet Potatoes"): ["1 large or 2 small", "2 large or 4 small", "3 large or 6 small", "4 large or 8 small", "5 large or 10 small", "6 large or 12 small"],
    ("sweet-potatoes", "Chickpeas"): ["3/4 cup", "1 1/2 cups", "2 1/4 cups", "3 cups", "3 3/4 cups", "4 1/2 cups"],
    ("sweet-potatoes", "Walnuts"): ["2 tbsp", "1/4 cup", "6 tbsp", "1/2 cup", "10 tbsp", "3/4 cup"],
    ("sweet-potatoes", "Coriander"): ["1 tbsp chopped", "2 tbsp chopped", "3 tbsp chopped", "1/4 cup chopped", "5 tbsp chopped", "3/8 cup chopped"],
    ("sweet-potatoes", "Mint"): ["1 tbsp chopped", "2 tbsp chopped", "3 tbsp chopped", "1/4 cup chopped", "5 tbsp chopped", "3/8 cup chopped"],
    ("sweet-potatoes", "Basil"): ["1 tbsp chopped", "2 tbsp chopped", "3 tbsp chopped", "1/4 cup chopped", "5 tbsp chopped", "3/8 cup chopped"],
    ("sweet-potatoes", "Thyme"): ["1 tsp leaves", "2 tsp leaves", "1 tbsp leaves", "4 tsp leaves", "5 tsp leaves", "2 tbsp leaves"],
    ("sweet-potatoes", "Olive Oil"): ["2 tsp", "4 tsp", "2 tbsp", "8 tsp", "10 tsp", "4 tbsp"],
    ("beans", "Green Beans"): ["1 1/2 cups", "3 cups", "4 1/2 cups", "6 cups", "7 1/2 cups", "9 cups"],
    ("beans", "Olive Oil"): ["1 tsp", "2 tsp", "1 tbsp", "4 tsp", "5 tsp", "2 tbsp"],
    ("cauliflower", "Cauliflower"): ["2 cups florets", "4 cups florets", "6 cups florets", "8 cups florets", "10 cups florets", "12 cups florets"],
    ("cauliflower", "Olive Oil"): ["1 tsp", "2 tsp", "1 tbsp", "4 tsp", "5 tsp", "2 tbsp"],
    ("ladys-finger", "Lady's Finger (Okra)"): ["1 1/2 cups", "3 cups", "4 1/2 cups", "6 cups", "7 1/2 cups", "9 cups"],
    ("ladys-finger", "Olive Oil"): ["1 tsp", "2 tsp", "1 tbsp", "4 tsp", "5 tsp", "2 tbsp"],
}

WEEKLY_BUY_PLANS = {
    "coconut water": ["2 to 3 liters", "4 to 5 liters", "6 to 7 liters", "8 to 9 liters", "10 to 11 liters", "12 to 13 liters"],
    "banana": ["7 bananas", "14 bananas", "21 bananas", "28 bananas", "35 bananas", "42 bananas"],
    "blueberries": ["3 cups", "5 cups", "7 cups", "9 cups", "11 cups", "13 cups"],
    "dragon fruit": ["3 dragon fruits", "5 dragon fruits", "7 dragon fruits", "9 dragon fruits", "11 dragon fruits", "13 dragon fruits"],
    "mango": ["4 mangoes", "7 mangoes", "10 mangoes", "13 mangoes", "16 mangoes", "19 mangoes"],
    "navel orange": ["4 oranges", "7 oranges", "10 oranges", "13 oranges", "16 oranges", "19 oranges"],
    "basil": ["1 bunch", "1 bunch", "2 bunches", "2 bunches", "3 bunches", "3 bunches"],
    "coriander": ["1 bunch", "1 bunch", "2 bunches", "2 bunches", "3 bunches", "3 bunches"],
    "mint": ["1 bunch", "1 bunch", "2 bunches", "2 bunches", "3 bunches", "3 bunches"],
    "thyme": ["1 small bunch", "1 small bunch", "1 small bunch", "2 small bunches", "2 small bunches", "2 small bunches"],
    "chickpeas": ["2 cans or 3 cups cooked", "3 cans or 4 1/2 cups cooked", "4 cans or 6 cups cooked", "5 cans or 7 1/2 cups cooked", "6 cans or 9 cups cooked", "7 cans or 10 1/2 cups cooked"],
    "coconut oil": ["1 small jar", "1 small jar", "1 small jar", "2 small jars", "2 small jars", "2 small jars"],
    "flax seeds": ["1 cup (about 250 g)", "2 cups (about 500 g)", "3 cups (about 750 g)", "4 cups (about 1 kg)", "5 cups (about 1.25 kg)", "6 cups (about 1.5 kg)"],
    "olive oil": ["1 bottle", "1 bottle", "1 bottle", "2 bottles", "2 bottles", "2 bottles"],
    "sesame seeds": ["1 small packet or jar", "1 small packet or jar", "1 small packet or jar", "1 small packet or jar", "2 small packets or jars", "2 small packets or jars"],
    "walnuts": ["1 1/2 cups", "3 cups", "4 1/2 cups", "6 cups", "7 1/2 cups", "9 cups"],
    "broccoli": ["4 medium heads", "7 medium heads", "10 medium heads", "13 medium heads", "16 medium heads", "19 medium heads"],
    "cauliflower": ["2 medium heads", "3 medium heads", "4 medium heads", "5 medium heads", "6 medium heads", "7 medium heads"],
    "celery": ["3 bunches", "5 bunches", "7 bunches", "9 bunches", "11 bunches", "13 bunches"],
    "green beans": ["1 1/2 lb", "3 lb", "4 1/2 lb", "6 lb", "7 1/2 lb", "9 lb"],
    "okra": ["1 1/2 lb", "3 lb", "4 1/2 lb", "6 lb", "7 1/2 lb", "9 lb"],
    "sweet potatoes": ["6 medium sweet potatoes", "10 medium sweet potatoes", "14 medium sweet potatoes", "18 medium sweet potatoes", "22 medium sweet potatoes", "26 medium sweet potatoes"],
    "chicken breast": ["2 lb", "3 1/2 lb", "5 lb", "6 1/2 lb", "8 lb", "9 1/2 lb"],
}

RECIPE_DETAILS = {
    "flax-seed-water": {
        "title": "Flax Seed Water",
        "summary": "A simple soaked drink that feels light but grounding first thing in the morning.",
        "method": "Soak and Stir",
        "prepMinutes": 2,
        "cookMinutes": 0,
        "cookGuide": ["No heat", "Soak 6 hr or overnight"],
        "ingredients": [
            {"name": "Flax Seeds", "amount": "1 tbsp"},
            {"name": "Water", "amount": "12 to 14 oz"},
        ],
        "steps": [
            "Stir 1 tablespoon flax seeds into a large glass of water.",
            "Let it soak overnight or at least 6 hours so the water softens and thickens slightly.",
            "Stir once more before drinking. Sip slowly while the seeds are suspended.",
        ],
        "tip": "Chilled water makes this easier to drink, and the brief soak softens the texture noticeably.",
    },
    "celery-juice": {
        "title": "Celery Juice",
        "summary": "Clean, fresh, and mild, with the option to keep the fiber in for a fuller drink.",
        "method": "Blend or Juice",
        "prepMinutes": 5,
        "cookMinutes": 0,
        "cookGuide": ["No heat", "Blend 1 to 2 min"],
        "ingredients": [
            {"name": "Celery", "amount": "4 to 5 stalks"},
            {"name": "Water", "amount": "2 to 3 tbsp if blending"},
        ],
        "steps": [
            "Roughly chop celery and add it to a blender with a small splash of water.",
            "Blend until smooth, adding just enough water to help it move.",
            "Drink as-is for more fiber, or strain for a lighter juice texture.",
        ],
        "tip": "Use only a little water so the celery flavor stays fresh instead of diluted.",
    },
    "mango-smoothie": {
        "title": "Mango Smoothie",
        "summary": "Thick, naturally sweet, and creamy from banana without needing anything extra.",
        "method": "Blender",
        "prepMinutes": 5,
        "cookMinutes": 0,
        "cookGuide": ["No heat", "Blend 45 to 60 sec"],
        "ingredients": [
            {"name": "Mango", "amount": "1 heaping cup"},
            {"name": "Banana", "amount": "1 medium"},
            {"name": "Coconut Water", "amount": "3/4 to 1 cup"},
        ],
        "steps": [
            "Add chopped mango, banana, and cold coconut water to a blender.",
            "Blend until silky and thick, scraping down the sides if needed.",
            "Pour immediately for the freshest taste and creamiest texture.",
        ],
        "tip": "If the fruit is cold, the smoothie tastes richer without needing ice.",
    },
    "berry-smoothie": {
        "title": "Berry Smoothie",
        "summary": "Bright and juicy, with dragon fruit keeping it light and blueberries adding depth.",
        "method": "Blender",
        "prepMinutes": 5,
        "cookMinutes": 0,
        "cookGuide": ["No heat", "Blend 45 to 60 sec"],
        "ingredients": [
            {"name": "Dragon Fruit", "amount": "1 cup"},
            {"name": "Blueberries", "amount": "3/4 cup"},
            {"name": "Banana", "amount": "1 medium"},
            {"name": "Coconut Water", "amount": "3/4 cup"},
        ],
        "steps": [
            "Add dragon fruit, blueberries, banana, and coconut water to a blender.",
            "Blend until the berries disappear and the color turns evenly vibrant.",
            "Taste for thickness and add a small splash of coconut water only if needed.",
        ],
        "tip": "Blend the blueberries fully so the drink stays smooth instead of slightly seedy.",
    },
    "orange-chicken": {
        "title": "Orange Chicken",
        "summary": "Citrusy and savory with caramelized edges on the chicken and crisp-tender broccoli.",
        "method": "Pan and Oven",
        "prepMinutes": 12,
        "cookMinutes": 20,
        "cookGuide": ["Pan: medium-high", "Oven: 425F", "Chicken: 165F inside"],
        "ingredients": [
            {"name": "Chicken Breast", "amount": "6 to 8 oz"},
            {"name": "Navel Orange", "amount": "1 whole"},
            {"name": "Broccoli", "amount": "1 1/2 to 2 cups florets"},
            {"name": "Coconut Oil", "amount": "1 tsp"},
            {"name": "Olive Oil", "amount": "1 tsp"},
            {"name": "Sesame Seeds", "amount": "1 tsp"},
        ],
        "steps": [
            "Preheat the oven to 425F. Pat the chicken dry, season both sides with salt and pepper, and heat a pan over medium-high heat for about 1 minute before adding the coconut oil.",
            "Sear the chicken for 3 to 4 minutes on the first side and 2 to 3 minutes on the second side, just until golden outside but not fully cooked through.",
            "Move the chicken to a small tray or oven-safe pan and bake at 425F for 8 to 12 minutes, or until the thickest part reaches 165F.",
            "While the chicken finishes, toss the broccoli with olive oil, salt, and pepper and roast it at 425F for 12 to 15 minutes until the edges darken and the stems feel tender.",
            "Set the chicken aside to rest for 3 minutes. Put the original pan back over medium heat, squeeze in the navel orange juice, and let it bubble for 2 to 3 minutes until it reduces slightly and looks glossy.",
            "Slice the chicken, coat it in the orange glaze, plate with the roasted broccoli, and finish with sesame seeds.",
        ],
        "tip": "If the orange juice starts boiling too hard, lower the pan to medium-low so it reduces gently instead of turning bitter.",
    },
    "sweet-potatoes": {
        "title": "Sweet Potatoes",
        "summary": "Roasted sweet potatoes layered with warm chickpeas, toasted walnuts, and a fresh herb finish.",
        "method": "Oven and Pan",
        "prepMinutes": 15,
        "cookMinutes": 30,
        "cookGuide": ["Oven: 425F", "Pan: medium", "Walnuts: 2 to 3 min"],
        "ingredients": [
            {"name": "Sweet Potatoes", "amount": "1 large or 2 small"},
            {"name": "Chickpeas", "amount": "3/4 cup"},
            {"name": "Walnuts", "amount": "2 tbsp"},
            {"name": "Coriander", "amount": "1 tbsp chopped"},
            {"name": "Mint", "amount": "1 tbsp chopped"},
            {"name": "Basil", "amount": "1 tbsp chopped"},
            {"name": "Thyme", "amount": "1 tsp leaves"},
            {"name": "Olive Oil", "amount": "2 tsp"},
        ],
        "steps": [
            "Preheat the oven to 425F. Cut the sweet potatoes into cubes or lengthwise halves, toss with olive oil, thyme, salt, and pepper, and spread them on a tray in one layer.",
            "Roast the sweet potatoes at 425F for 25 to 30 minutes, flipping once halfway through, until the edges are browned and the center is soft when pierced with a fork.",
            "While they roast, heat a pan over medium heat with a small splash of olive oil and cook the chickpeas for 5 to 6 minutes, stirring occasionally, until they feel warmer, drier, and slightly crisp.",
            "In a dry pan over medium-low heat, toast the walnuts for 2 to 3 minutes, shaking the pan often, until fragrant. Take them off the heat as soon as they smell nutty.",
            "Finely chop the coriander, mint, and basil together so they are ready as a fresh topping.",
            "Pile the warm chickpeas and walnuts over the roasted sweet potatoes and finish with the chopped herbs right before serving.",
        ],
        "tip": "Keep the herbs off the heat until the end so the bowl tastes fresh instead of muted.",
    },
    "beans": {
        "title": "Beans",
        "summary": "A fast pan-cooked green bean side with blistered spots and a clean savory finish.",
        "method": "Stovetop Pan",
        "prepMinutes": 5,
        "cookMinutes": 10,
        "cookGuide": ["Pan: medium-high", "Cook: 6 to 8 min"],
        "ingredients": [
            {"name": "Green Beans", "amount": "1 1/2 cups"},
            {"name": "Olive Oil", "amount": "1 tsp"},
        ],
        "steps": [
            "Heat a wide pan over medium-high heat for about 1 minute, then add the olive oil.",
            "Add the green beans in one layer, season with salt and pepper, and let them sit undisturbed for 2 minutes so they start to blister.",
            "Continue cooking for another 4 to 6 minutes, tossing only once or twice, until the beans have browned spots but still keep a slight snap.",
            "Serve right away while the beans still have a little snap.",
        ],
        "tip": "Leaving the beans mostly undisturbed at first gives better browning and more flavor.",
    },
    "cauliflower": {
        "title": "Cauliflower",
        "summary": "Golden roasted cauliflower that turns sweet and nutty without needing extra seasoning.",
        "method": "Oven or Air Fryer",
        "prepMinutes": 8,
        "cookMinutes": 18,
        "cookGuide": ["Oven: 425F", "Air fryer: 390F", "Cook: 12 to 22 min"],
        "ingredients": [
            {"name": "Cauliflower", "amount": "2 cups florets"},
            {"name": "Olive Oil", "amount": "1 tsp"},
        ],
        "steps": [
            "For the oven method, preheat to 425F. For the air fryer method, preheat to 390F for 3 minutes if your model supports it.",
            "Cut the cauliflower into medium florets and coat with olive oil, salt, and pepper.",
            "Roast in the oven at 425F for 18 to 22 minutes, or air fry at 390F for 12 to 15 minutes, turning once halfway through in either method.",
            "Cook until the edges are deeply golden and the thickest pieces are tender when pierced with a fork.",
            "Serve hot while the edges are still crisp.",
        ],
        "tip": "Do not overcrowd the tray or basket or the cauliflower will steam instead of roast.",
    },
    "ladys-finger": {
        "title": "Lady's Finger",
        "summary": "Dry-cooked okra with lightly charred edges and none of the slimy texture people worry about.",
        "method": "Stovetop Pan",
        "prepMinutes": 8,
        "cookMinutes": 12,
        "cookGuide": ["Pan: medium-high", "Cook: 8 to 10 min"],
        "ingredients": [
            {"name": "Lady's Finger (Okra)", "amount": "1 1/2 cups"},
            {"name": "Olive Oil", "amount": "1 tsp"},
        ],
        "steps": [
            "Dry the okra very well with a towel, then slice it into larger pieces so it browns instead of steaming.",
            "Heat a wide pan over medium-high heat for about 1 minute, add the olive oil, then spread the okra in one layer.",
            "Season with salt and pepper and cook for 3 minutes without stirring so the first side starts to char.",
            "Cook for another 5 to 7 minutes, stirring only occasionally, until the cut edges darken and the pieces feel dry and tender.",
        ],
        "tip": "If the okra starts releasing too much moisture, lower the heat slightly and keep cooking uncovered so the pan dries back out.",
    },
}

WEEKLY_QUANTITY_GUIDE = {
    "flax seeds": {"quantity": "1 cup (about 250 g)", "category": "Pantry"},
    "celery": {"quantity": "3 bunches", "category": "Produce"},
    "mango": {"quantity": "4 mangoes", "category": "Fruit"},
    "banana": {"quantity": "7 bananas", "category": "Fruit"},
    "coconut water": {"quantity": "2 to 3 liters", "category": "Beverages"},
    "dragon fruit": {"quantity": "3 dragon fruits", "category": "Fruit"},
    "blueberries": {"quantity": "3 cups", "category": "Fruit"},
    "chicken breast": {"quantity": "2 lb", "category": "Protein"},
    "navel orange": {"quantity": "4 oranges", "category": "Fruit"},
    "broccoli": {"quantity": "4 medium heads", "category": "Produce"},
    "sesame seeds": {"quantity": "1 small packet or jar", "category": "Pantry"},
    "sweet potatoes": {"quantity": "6 medium sweet potatoes", "category": "Produce"},
    "chickpeas": {"quantity": "2 cans or 3 cups cooked", "category": "Pantry"},
    "walnuts": {"quantity": "1 1/2 cups", "category": "Pantry"},
    "coriander": {"quantity": "1 bunch", "category": "Herbs"},
    "mint": {"quantity": "1 bunch", "category": "Herbs"},
    "basil": {"quantity": "1 bunch", "category": "Herbs"},
    "thyme": {"quantity": "1 small bunch", "category": "Herbs"},
    "green beans": {"quantity": "1 1/2 lb", "category": "Produce"},
    "cauliflower": {"quantity": "2 medium heads", "category": "Produce"},
    "okra": {"quantity": "1 1/2 lb", "category": "Produce"},
    "olive oil": {"quantity": "1 bottle", "category": "Pantry"},
    "coconut oil": {"quantity": "1 small jar", "category": "Pantry"},
}

DEFAULT_STORE = "Whole Foods / Fred Meyer / QFC"


def build_plan_map(values: list[str]) -> dict[str, str]:
    return {str(index + 1): value for index, value in enumerate(values)}


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "item"


def dish_id_from_label(label: str) -> str:
    recipe_key = canonicalize(label)
    recipe_slug_map = {
        "flax seeds": "flax-seed-water",
        "celery": "celery-juice",
        "green beans": "beans",
        "okra": "ladys-finger",
    }
    return recipe_slug_map.get(recipe_key, slugify(label))


def canonicalize(name: str) -> str:
    normalized = re.sub(r"[^a-z0-9' ]+", "", name.lower()).strip()
    return CANONICAL_OVERRIDES.get(normalized, normalized)


def display_name(name: str) -> str:
    if name in DISPLAY_NAMES:
        return DISPLAY_NAMES[name]
    return " ".join(part.capitalize() for part in name.split())


def load_sheet():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    return workbook["Sheet1"]


def parse_meal_plan(ws) -> list[dict]:
    meal_rows = [
        ("Morning Cleanse", 7, 8),
        ("Breakfast", 10),
        ("Lunch", 12),
        ("Dinner", 14),
    ]
    days = []
    for index, col in enumerate(range(4, 11)):
        meals = []
        for row_info in meal_rows:
            name = row_info[0]
            rows = row_info[1:]
            items = [clean(ws.cell(row, col).value) for row in rows if clean(ws.cell(row, col).value)]
            meals.append(
                {
                    "mealType": name,
                    "dishIds": [dish_id_from_label(item) for item in items],
                    "labels": items,
                }
            )
        days.append(
            {
                "id": DAY_NAMES[index].lower(),
                "name": DAY_NAMES[index],
                "meals": meals,
            }
        )
    return days


def parse_workbook_recipe_blocks(ws) -> dict[str, list[str]]:
    recipes = {}
    for col in range(13, ws.max_column + 1):
        row = 1
        while row <= ws.max_row:
            current = clean(ws.cell(row, col).value)
            previous = clean(ws.cell(row - 1, col).value) if row > 1 else ""
            next_value = clean(ws.cell(row + 1, col).value) if row < ws.max_row else ""
            if current and not previous and next_value:
                items = []
                cursor = row + 1
                while cursor <= ws.max_row:
                    ingredient = clean(ws.cell(cursor, col).value)
                    if not ingredient:
                        break
                    items.append(ingredient)
                    cursor += 1
                recipes[current] = items
                row = cursor
            else:
                row += 1
    return recipes


def parse_store_lists(ws) -> list[dict]:
    stores = []
    for col in range(4, ws.max_column + 1):
        store = clean(ws.cell(23, col).value)
        if not store:
            continue
        items = []
        row = 24
        while row <= ws.max_row:
            item = clean(ws.cell(row, col).value)
            if not item:
                break
            canonical = canonicalize(item)
            items.append(
                {
                    "name": display_name(canonical),
                    "canonical": canonical,
                    "source": item,
                }
            )
            row += 1
        stores.append({"store": store, "items": items})
    return stores


def build_recipes(workbook_recipes: dict[str, list[str]]) -> tuple[list[dict], dict[str, dict]]:
    derived = []
    recipe_by_id = {}
    for recipe_id, details in RECIPE_DETAILS.items():
        workbook_ingredients = workbook_recipes.get(details["title"], [])
        listed_ingredients = []
        for ingredient in details["ingredients"]:
            canonical = canonicalize(ingredient["name"])
            listed_ingredients.append(
                {
                    "name": display_name(canonical),
                    "canonical": canonical,
                    "amount": ingredient["amount"],
                    "amountPlan": build_plan_map(RECIPE_AMOUNT_PLANS[(recipe_id, ingredient["name"])]),
                }
            )
        recipe = {
            "id": recipe_id,
            "title": details["title"],
            "summary": details["summary"],
            "method": details["method"],
            "icon": DISH_VISUALS.get(recipe_id, {}).get("icon", "🍽️"),
            "avatar": DISH_VISUALS.get(recipe_id, {}).get("avatar", "PL"),
            "prepMinutes": details["prepMinutes"],
            "cookMinutes": details["cookMinutes"],
            "cookGuide": details.get("cookGuide", []),
            "ingredients": listed_ingredients,
            "steps": details["steps"],
            "tip": details["tip"],
            "workbookIngredients": workbook_ingredients,
        }
        derived.append(recipe)
        recipe_by_id[recipe_id] = recipe
    return derived, recipe_by_id


def dish_recipe_id(dish_label: str) -> str:
    return dish_id_from_label(dish_label)


def build_dishes(days: list[dict], recipe_by_id: dict[str, dict]) -> list[dict]:
    seen = {}
    for day in days:
        for meal in day["meals"]:
            for label in meal["labels"]:
                dish_id = slugify(label)
                dish_id = dish_id_from_label(label)
                if dish_id in seen:
                    continue
                recipe = recipe_by_id[dish_recipe_id(label)]
                seen[dish_id] = {
                    "id": dish_id,
                    "title": label,
                    "recipeId": recipe["id"],
                    "summary": recipe["summary"],
                    "icon": recipe["icon"],
                    "avatar": recipe["avatar"],
                }
    return list(seen.values())


def build_store_lookup(store_lists: list[dict]) -> dict[str, list[str]]:
    lookup = {}
    for store in store_lists:
        for item in store["items"]:
            lookup.setdefault(item["canonical"], []).append(store["store"])
    return lookup


def build_weekly_supply_list(recipe_by_id: dict[str, dict], store_lookup: dict[str, list[str]]) -> list[dict]:
    ingredient_keys = []
    for recipe in recipe_by_id.values():
        for ingredient in recipe["ingredients"]:
            canonical = ingredient["canonical"]
            if canonical in {"water"}:
                continue
            if canonical not in ingredient_keys:
                ingredient_keys.append(canonical)

    grouped = {}
    for canonical in ingredient_keys:
        guide = WEEKLY_QUANTITY_GUIDE.get(canonical)
        if not guide:
            continue
        store = store_lookup.get(canonical, [DEFAULT_STORE])[0]
        grouped.setdefault(guide["category"], []).append(
            {
                "name": display_name(canonical),
                "quantity": guide["quantity"],
                "buyPlan": build_plan_map(WEEKLY_BUY_PLANS[canonical]),
                "store": store,
            }
        )

    return [
        {
            "category": category,
            "items": sorted(items, key=lambda item: item["name"].lower()),
        }
        for category, items in sorted(grouped.items(), key=lambda pair: pair[0].lower())
    ]


def build_daily_supply_list(days: list[dict], dishes: list[dict], recipe_by_id: dict[str, dict], store_lookup: dict[str, list[str]]) -> dict[str, list[dict]]:
    dish_lookup = {dish["id"]: dish for dish in dishes}
    output = {}
    for day in days:
        grouped = {}
        seen = set()
        for meal in day["meals"]:
            for dish_id in meal["dishIds"]:
                recipe = recipe_by_id[dish_lookup[dish_id]["recipeId"]]
                for ingredient in recipe["ingredients"]:
                    key = (ingredient["canonical"], ingredient["amount"])
                    if key in seen or ingredient["canonical"] == "water":
                        continue
                    seen.add(key)
                    guide = WEEKLY_QUANTITY_GUIDE.get(ingredient["canonical"])
                    category = guide["category"] if guide else "Needs Review"
                    store = store_lookup.get(ingredient["canonical"], [DEFAULT_STORE])[0]
                    grouped.setdefault(category, []).append(
                        {
                            "name": ingredient["name"],
                            "used": ingredient["amount"],
                            "usedPlan": ingredient["amountPlan"],
                            "buy": guide["quantity"] if guide else "See weekly supply list",
                            "buyPlan": build_plan_map(WEEKLY_BUY_PLANS[ingredient["canonical"]]) if ingredient["canonical"] in WEEKLY_BUY_PLANS else build_plan_map(["See weekly supply list"] * 6),
                            "store": store,
                        }
                    )
        output[day["id"]] = [
            {"category": category, "items": sorted(items, key=lambda item: item["name"].lower())}
            for category, items in sorted(grouped.items(), key=lambda pair: pair[0].lower())
        ]
    return output


def build_day_shopping(days: list[dict], recipe_by_id: dict[str, dict], store_lookup: dict[str, list[str]]) -> dict[str, dict]:
    result = {}
    all_unassigned = set()
    for day in days:
        store_map = {}
        unassigned = set()
        for meal in day["meals"]:
            for dish_id in meal["dishIds"]:
                recipe = recipe_by_id[dish_id]
                for ingredient in recipe["ingredients"]:
                    canonical = ingredient["canonical"]
                    if canonical in {"water", "salt", "black pepper", "neutral oil", "olive oil"}:
                        continue
                    stores = store_lookup.get(canonical, [])
                    if stores:
                        for store in stores:
                            store_map.setdefault(store, set()).add(ingredient["name"])
                    else:
                        unassigned.add(ingredient["name"])
                        all_unassigned.add(ingredient["name"])
        result[day["id"]] = {
            "stores": [
                {"store": store, "items": sorted(items)}
                for store, items in sorted(store_map.items(), key=lambda pair: pair[0].lower())
            ],
            "unassigned": sorted(unassigned),
        }
    result["weekly"] = {"unassigned": sorted(all_unassigned)}
    return result


def build_weekly_shopping(store_lists: list[dict], all_unassigned: list[str]) -> list[dict]:
    weekly = []
    for store in store_lists:
        weekly.append(
            {
                "store": store["store"],
                "items": [item["name"] for item in store["items"]],
            }
        )
    if all_unassigned:
        weekly.append({"store": "Needs Store Assignment", "items": all_unassigned})
    return weekly


def build_data() -> dict:
    ws = load_sheet()
    days = parse_meal_plan(ws)
    workbook_recipes = parse_workbook_recipe_blocks(ws)
    store_lists = parse_store_lists(ws)
    recipes, recipe_by_id = build_recipes(workbook_recipes)
    dishes = build_dishes(days, recipe_by_id)
    store_lookup = build_store_lookup(store_lists)
    day_shopping = build_day_shopping(days, recipe_by_id, store_lookup)
    weekly_shopping = build_weekly_shopping(store_lists, day_shopping["weekly"]["unassigned"])
    weekly_supply = build_weekly_supply_list(recipe_by_id, store_lookup)
    daily_supply = build_daily_supply_list(days, dishes, recipe_by_id, store_lookup)

    return {
        "title": "Weekly Diet Dashboard",
        "subtitle": "Tap a day to see the meals, recipes, and shopping details that matter right then.",
        "generatedAt": datetime.now().strftime("%B %d, %Y"),
        "defaultDay": "monday",
        "pantryBasics": PANTRY_BASICS,
        "days": days,
        "dishes": dishes,
        "recipes": recipes,
        "dayShopping": day_shopping,
        "dailySupply": daily_supply,
        "weeklySupply": weekly_supply,
        "weeklyShopping": weekly_shopping,
        "pdfPath": "output/weekly-diet-dashboard.pdf",
    }


def build_html(data: dict) -> str:
    title = escape(data["title"])
    subtitle = escape(data["subtitle"])
    embedded_json = json.dumps(data, indent=2)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{title}</title>
  <meta name="description" content="{escape(data['subtitle'])}" />
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700;800&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg-1: #fff0de;
      --bg-2: #ffe4f2;
      --bg-3: #e5fff1;
      --panel: rgba(255, 250, 247, 0.88);
      --panel-strong: rgba(255, 255, 255, 0.92);
      --ink: #182329;
      --muted: #56616a;
      --accent: #ff5c8a;
      --accent-strong: #ff3f74;
      --accent-soft: #ffe3ee;
      --mint: #7ce7b8;
      --sky: #c9f4ff;
      --warm: #ff9b42;
      --line: rgba(24, 35, 41, 0.08);
      --shadow: 0 24px 48px rgba(100, 67, 95, 0.12);
      --radius-xl: 28px;
      --radius-lg: 22px;
      --radius-md: 16px;
      --max: 1120px;
    }}
    * {{ box-sizing: border-box; }}
    html {{ scroll-behavior: smooth; }}
    body {{
      margin: 0;
      color: var(--ink);
      font-family: "DM Sans", "Segoe UI", sans-serif;
      background:
        radial-gradient(circle at top left, rgba(255, 92, 138, 0.18), transparent 22%),
        radial-gradient(circle at top right, rgba(124, 231, 184, 0.24), transparent 25%),
        radial-gradient(circle at bottom left, rgba(201, 244, 255, 0.24), transparent 20%),
        linear-gradient(180deg, var(--bg-1) 0%, #fff7ef 42%, var(--bg-2) 78%, var(--bg-3) 100%);
      min-height: 100vh;
    }}
    .shell {{
      width: min(calc(100% - 20px), var(--max));
      margin: 0 auto;
      padding: 18px 0 40px;
    }}
    .hero {{
      position: relative;
      overflow: hidden;
      border-radius: 32px;
      padding: 24px;
      background:
        linear-gradient(135deg, rgba(255, 255, 255, 0.95), rgba(255, 244, 249, 0.82)),
        radial-gradient(circle at top right, rgba(124, 231, 184, 0.24), transparent 35%);
      border: 1px solid rgba(255, 255, 255, 0.75);
      box-shadow: var(--shadow);
    }}
    .hero::after {{
      content: "";
      position: absolute;
      width: 260px;
      height: 260px;
      right: -60px;
      bottom: -120px;
      border-radius: 50%;
      background: radial-gradient(circle, rgba(255, 92, 138, 0.18), transparent 68%);
      pointer-events: none;
    }}
    @keyframes floaty {{
      0%, 100% {{ transform: translateY(0px); }}
      50% {{ transform: translateY(-4px); }}
    }}
    @keyframes fadeRise {{
      from {{ opacity: 0; transform: translateY(12px); }}
      to {{ opacity: 1; transform: translateY(0); }}
    }}
    .eyebrow {{
      display: inline-flex;
      gap: 8px;
      align-items: center;
      padding: 7px 12px;
      border-radius: 999px;
      background: rgba(255, 92, 138, 0.1);
      color: var(--accent-strong);
      font-size: 12px;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      font-weight: 700;
    }}
    h1 {{
      margin: 14px 0 10px;
      font-family: "Space Grotesk", "DM Sans", sans-serif;
      font-size: clamp(30px, 7vw, 60px);
      line-height: 0.95;
      max-width: 9ch;
      letter-spacing: -0.04em;
    }}
    .subtitle {{
      max-width: 52ch;
      margin: 0;
      color: var(--muted);
      font-size: 15px;
      line-height: 1.55;
    }}
    .hero-meta {{
      margin-top: 18px;
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}
    .meta-pill, .method-pill, .store-pill, .pantry-pill {{
      display: inline-flex;
      align-items: center;
      padding: 8px 12px;
      border-radius: 999px;
      border: 1px solid rgba(32, 49, 39, 0.08);
      background: rgba(255, 255, 255, 0.78);
      font-size: 12px;
      color: var(--muted);
    }}
    .top-actions {{
      margin-top: 16px;
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}
    .action-link {{
      border: 0;
      text-decoration: none;
      background: linear-gradient(135deg, var(--accent), var(--accent-strong));
      color: white;
      border-radius: 999px;
      padding: 12px 16px;
      font-weight: 700;
      box-shadow: 0 12px 26px rgba(255, 92, 138, 0.24);
      cursor: pointer;
      font: inherit;
    }}
    .action-link.secondary {{
      background: rgba(255, 255, 255, 0.86);
      color: var(--ink);
      box-shadow: none;
    }}
    .day-bar {{
      position: sticky;
      top: 0;
      z-index: 5;
      margin: 18px 0 22px;
      padding: 10px 0 4px;
      backdrop-filter: blur(16px);
    }}
    .day-track {{
      display: flex;
      gap: 10px;
      overflow-x: auto;
      padding-bottom: 6px;
      scrollbar-width: none;
    }}
    .day-track::-webkit-scrollbar {{ display: none; }}
    .day-chip {{
      flex: 0 0 auto;
      border: 0;
      border-radius: 999px;
      padding: 13px 16px;
      background: rgba(255, 255, 255, 0.76);
      color: var(--ink);
      font-weight: 800;
      font-size: 13px;
      box-shadow: 0 10px 20px rgba(100, 67, 95, 0.08);
      cursor: pointer;
    }}
    .day-chip.active {{
      background: linear-gradient(135deg, var(--accent), var(--accent-strong));
      color: white;
      box-shadow: 0 14px 26px rgba(255, 92, 138, 0.28);
    }}
    .focus-layout, .logistics-layout {{
      display: grid;
      gap: 18px;
    }}
    .recipe-stage, .pantry-stage {{
      min-width: 0;
    }}
    .section-heading {{
      display: flex;
      justify-content: space-between;
      align-items: end;
      gap: 16px;
      margin: 22px 0 14px;
    }}
    .section-heading.compact {{
      margin-top: 0;
    }}
    .people-control {{
      display: inline-flex;
      align-items: center;
      gap: 10px;
      padding: 10px 12px;
      border-radius: 16px;
      background: rgba(255,255,255,0.76);
      border: 1px solid rgba(24, 35, 41, 0.08);
    }}
    .people-control span {{
      color: var(--muted);
      font-size: 13px;
      font-weight: 700;
    }}
    .people-stepper {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
    }}
    .people-stepper button {{
      width: 32px;
      height: 32px;
      border: 0;
      border-radius: 999px;
      background: linear-gradient(135deg, var(--accent), var(--accent-strong));
      color: white;
      font: inherit;
      font-weight: 800;
      cursor: pointer;
      box-shadow: 0 8px 16px rgba(255, 92, 138, 0.18);
    }}
    .people-stepper strong {{
      min-width: 18px;
      text-align: center;
      font-size: 16px;
    }}
    .section-heading h2 {{
      margin: 0;
      font-family: "Space Grotesk", "DM Sans", sans-serif;
      font-size: 24px;
      letter-spacing: -0.03em;
    }}
    .section-heading p {{
      margin: 6px 0 0;
      color: var(--muted);
      line-height: 1.45;
      max-width: 58ch;
    }}
    .selected-day-shell, .panel {{
      background: var(--panel);
      border: 1px solid rgba(255, 255, 255, 0.7);
      box-shadow: var(--shadow);
      border-radius: var(--radius-xl);
    }}
    .recipe-stage .panel, .selected-day-shell, .shopping-card, .pantry-stage .panel {{
      animation: fadeRise 240ms ease;
    }}
    .selected-day-shell {{
      padding: 18px;
    }}
    .selected-day-head {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: end;
      margin-bottom: 16px;
    }}
    .selected-day-head h3 {{
      margin: 0;
      font-family: "Space Grotesk", "DM Sans", sans-serif;
      font-size: clamp(26px, 5vw, 40px);
      letter-spacing: -0.04em;
    }}
    .selected-day-head small {{
      display: block;
      margin-bottom: 6px;
      color: var(--warm);
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.08em;
    }}
    .meals-grid, .recipes-grid, .shopping-grid {{
      display: grid;
      gap: 16px;
    }}
    .meals-grid {{
      grid-template-columns: repeat(2, minmax(0, 1fr));
    }}
    .meal-card, .panel {{
      padding: 16px;
    }}
    .meal-card {{
      border-radius: var(--radius-lg);
      background: rgba(255, 255, 255, 0.7);
      border: 1px solid var(--line);
    }}
    .meal-card h4 {{
      margin: 0 0 10px;
      color: var(--muted);
      font-size: 13px;
      letter-spacing: 0.08em;
      text-transform: uppercase;
    }}
    .meal-list {{
      display: grid;
      gap: 10px;
    }}
    .meal-topline {{
      display: flex;
      gap: 10px;
      align-items: center;
    }}
    .dish-avatar {{
      width: 40px;
      height: 40px;
      border-radius: 14px;
      background: linear-gradient(135deg, rgba(255, 92, 138, 0.16), rgba(124, 231, 184, 0.28));
      display: grid;
      place-items: center;
      font-weight: 800;
      font-size: 12px;
      color: var(--ink);
      flex: 0 0 auto;
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.7);
      animation: floaty 3.2s ease-in-out infinite;
    }}
    .dish-icon {{
      font-size: 20px;
      line-height: 1;
    }}
    .meal-item {{
      display: block;
      text-decoration: none;
      color: inherit;
      width: 100%;
      text-align: left;
      padding: 12px;
      border-radius: 14px;
      background: var(--panel-strong);
      border: 1px solid rgba(32, 49, 39, 0.06);
      transition: transform 120ms ease, box-shadow 120ms ease, border-color 120ms ease;
    }}
    .meal-item:hover, .meal-item:focus {{
      transform: translateY(-1px);
      box-shadow: 0 10px 18px rgba(32, 49, 39, 0.08);
      border-color: rgba(31, 106, 82, 0.28);
      outline: none;
    }}
    .meal-item strong {{
      display: block;
      font-size: 15px;
      margin-bottom: 2px;
    }}
    .meal-item span {{
      color: var(--muted);
      font-size: 13px;
      line-height: 1.4;
    }}
    .recipes-grid, .shopping-grid {{
      grid-template-columns: repeat(2, minmax(0, 1fr));
    }}
    .panel h3 {{
      margin: 0 0 10px;
      font-family: "Space Grotesk", "DM Sans", sans-serif;
      font-size: 20px;
      letter-spacing: -0.03em;
    }}
    .recipe-card.flash {{
      outline: 3px solid rgba(31, 106, 82, 0.3);
      box-shadow: 0 0 0 6px rgba(31, 106, 82, 0.08);
    }}
    .panel p {{
      margin: 0;
      color: var(--muted);
      line-height: 1.5;
    }}
    .panel-top {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: start;
      margin-bottom: 12px;
    }}
    .recipe-title-group {{
      display: flex;
      gap: 12px;
      align-items: flex-start;
    }}
    .recipe-avatar {{
      width: 48px;
      height: 48px;
      border-radius: 18px;
      background: linear-gradient(135deg, rgba(255, 92, 138, 0.16), rgba(124, 231, 184, 0.28));
      display: grid;
      place-items: center;
      flex: 0 0 auto;
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.76);
      animation: floaty 3.6s ease-in-out infinite;
    }}
    .recipe-avatar .dish-icon {{
      font-size: 24px;
    }}
    .recipe-meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 12px 0;
    }}
    .cook-guide {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 0 0 14px;
    }}
    .guide-pill {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 8px 11px;
      border-radius: 999px;
      background: rgba(20, 36, 51, 0.08);
      border: 1px solid rgba(20, 36, 51, 0.08);
      color: #244258;
      font-size: 12px;
      font-weight: 700;
      letter-spacing: 0.01em;
    }}
    .ingredient-list, .steps-list, .shopping-list, .pantry-list {{
      margin: 0;
      padding: 0;
      list-style: none;
    }}
    .ingredient-list {{
      display: grid;
      gap: 8px;
      margin-bottom: 14px;
    }}
    .ingredient-list li {{
      background: linear-gradient(135deg, rgba(255, 227, 238, 0.9), rgba(201, 244, 255, 0.85));
      border-radius: 16px;
      padding: 10px 12px;
      font-size: 13px;
      border: 1px solid rgba(24, 35, 41, 0.06);
    }}
    .ingredient-list li strong {{
      display: block;
      font-size: 14px;
      margin-bottom: 3px;
    }}
    .steps-list {{
      display: grid;
      gap: 10px;
      counter-reset: recipe-step;
      margin-bottom: 12px;
    }}
    .steps-list li {{
      position: relative;
      padding-left: 38px;
      line-height: 1.45;
      color: var(--ink);
    }}
    .steps-list li::before {{
      counter-increment: recipe-step;
      content: counter(recipe-step);
      position: absolute;
      left: 0;
      top: 0;
      width: 26px;
      height: 26px;
      border-radius: 50%;
      background: rgba(31, 106, 82, 0.12);
      color: var(--accent-strong);
      font-size: 13px;
      font-weight: 800;
      display: grid;
      place-items: center;
    }}
    .tip {{
      margin-top: 10px;
      padding: 12px;
      border-radius: 14px;
      background: linear-gradient(135deg, rgba(255, 155, 66, 0.12), rgba(255, 92, 138, 0.08));
      color: #8a4c1a;
      font-size: 14px;
      line-height: 1.45;
    }}
    .section-sticker {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      padding: 6px 10px;
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.72);
      font-size: 12px;
      font-weight: 700;
      color: var(--muted);
      margin-bottom: 8px;
    }}
    .shopping-card h3 {{
      margin-bottom: 8px;
    }}
    .shopping-toggle-row {{
      margin: 0 0 14px;
    }}
    .toggle-button {{
      border: 0;
      border-radius: 999px;
      padding: 12px 16px;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      background: linear-gradient(135deg, var(--accent), var(--accent-strong));
      color: white;
      box-shadow: 0 10px 20px rgba(255, 92, 138, 0.2);
    }}
    .toggle-button.subtle {{
      background: rgba(255,255,255,0.82);
      color: var(--ink);
      box-shadow: none;
      border: 1px solid var(--line);
    }}
    .collapsible-section[hidden] {{
      display: none !important;
    }}
    .collapsible-section {{
      margin-bottom: 10px;
    }}
    .shopping-list {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 10px;
    }}
    .shopping-list li {{
      padding: 8px 12px;
      background: rgba(255, 255, 255, 0.8);
      border: 1px solid rgba(32, 49, 39, 0.08);
      border-radius: 999px;
      font-size: 14px;
    }}
    .shopping-note {{
      margin-top: 12px;
      color: var(--muted);
      line-height: 1.45;
    }}
    .weekly-supply-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 16px;
    }}
    .supply-list {{
      list-style: none;
      margin: 12px 0 0;
      padding: 0;
      display: grid;
      gap: 10px;
    }}
    .supply-list li {{
      padding: 12px;
      border-radius: 14px;
      background: rgba(255, 255, 255, 0.78);
      border: 1px solid rgba(32, 49, 39, 0.08);
    }}
    .supply-list strong {{
      display: block;
      margin-bottom: 4px;
    }}
    .supply-list span {{
      display: block;
      color: var(--muted);
      line-height: 1.4;
      font-size: 13px;
    }}
    .buy-quantity {{
      color: var(--accent-strong);
      font-weight: 800;
      font-size: 15px;
      margin: 3px 0;
    }}
    .used-quantity {{
      color: var(--ink);
      font-weight: 700;
      font-size: 14px;
      margin: 3px 0;
    }}
    details.weekly {{
      margin-top: 18px;
      background: rgba(255, 255, 255, 0.52);
      border: 1px solid rgba(32, 49, 39, 0.08);
      border-radius: 20px;
      overflow: hidden;
    }}
    details.weekly summary {{
      cursor: pointer;
      list-style: none;
      padding: 16px 18px;
      font-weight: 800;
    }}
    details.weekly summary::-webkit-details-marker {{ display: none; }}
    .weekly-body {{
      padding: 0 18px 18px;
    }}
    .pantry-list {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 12px;
    }}
    .footer-note {{
      margin-top: 26px;
      color: var(--muted);
      text-align: center;
      font-size: 13px;
    }}
    @media (max-width: 880px) {{
      .meals-grid, .recipes-grid, .shopping-grid, .weekly-supply-grid {{
        grid-template-columns: 1fr;
      }}
      .selected-day-head {{
        flex-direction: column;
        align-items: start;
      }}
    }}
    @media (min-width: 881px) {{
      .focus-layout {{
        grid-template-columns: minmax(0, 1.08fr) minmax(360px, 0.92fr);
        align-items: start;
      }}
      .recipe-stage {{
        position: sticky;
        top: 88px;
      }}
      .logistics-layout {{
        grid-template-columns: repeat(2, minmax(0, 1fr));
      }}
      .pantry-stage {{
        grid-column: 1 / -1;
      }}
    }}
    @media (min-width: 1240px) {{
      .shell {{
        width: min(calc(100% - 36px), 1320px);
      }}
      .logistics-layout {{
        grid-template-columns: 1fr 1fr 0.78fr;
        align-items: start;
      }}
      .pantry-stage {{
        grid-column: auto;
      }}
    }}
    @media print {{
      @page {{
        size: A4;
        margin: 10mm;
      }}
      body {{
        background: white;
      }}
      .shell {{
        width: 100%;
        padding: 0;
      }}
      .hero, .selected-day-shell, .panel {{
        box-shadow: none;
        background: white;
      }}
      .day-bar {{
        position: static;
        backdrop-filter: none;
      }}
      .meals-grid, .recipes-grid, .shopping-grid, .weekly-supply-grid {{
        grid-template-columns: 1fr;
      }}
      .action-link {{
        color: var(--ink);
        border: 1px solid var(--line);
        background: white;
        box-shadow: none;
      }}
    }}
  </style>
</head>
<body>
  <main class="shell">
    <section class="hero">
      <span class="eyebrow">Public Share Link Ready</span>
      <h1>{title}</h1>
      <p class="subtitle">{subtitle}</p>
      <div class="top-actions">
        <a class="action-link" href="#day-view">Jump to Day View</a>
        <button class="action-link secondary" type="button" id="open-day-shopping">Day Shopping</button>
        <button class="action-link secondary" type="button" id="open-weekly-shopping">Weekly Shopping</button>
        <a class="action-link secondary" href="{escape(data['pdfPath'])}">Open PDF Backup</a>
      </div>
    </section>

    <nav class="day-bar" aria-label="Pick a day">
      <div class="day-track" id="day-track"></div>
    </nav>

    <section class="focus-layout">
      <section id="day-view" class="selected-day-shell">
        <div class="selected-day-head">
          <div>
            <small>Selected Day</small>
            <h3 id="selected-day-title">Monday</h3>
          </div>
          <p id="selected-day-copy" class="subtitle">Meals, recipes, and shopping are focused on the day you choose.</p>
        </div>
        <div class="meals-grid" id="meals-grid"></div>
      </section>

      <section class="recipe-stage">
        <div class="section-heading compact">
          <div>
            <div class="section-sticker">✨ Tap a meal to open its recipe</div>
            <h2>Recipe Focus</h2>
            <p>Only one recipe opens at a time so the page stays calm and easy to scan.</p>
          </div>
          <div class="people-control" id="recipe-people-control">
            <span>Recipe people</span>
            <div class="people-stepper">
              <button type="button" id="recipe-minus">-</button>
              <strong id="recipe-people-count">1</strong>
              <button type="button" id="recipe-plus">+</button>
            </div>
          </div>
        </div>
        <div id="recipe-empty" class="panel">
          <h3>Choose a meal card above</h3>
          <p>Tap any breakfast, lunch, dinner, or cleanse item to open the matching recipe with the exact amount used.</p>
        </div>
        <div class="recipes-grid" id="recipes-grid"></div>
      </section>
    </section>

    <section class="logistics-layout">
      <section>
        <div class="section-heading">
          <div>
            <div class="section-sticker">🛒 One-person weekly buy guide</div>
            <h2>Weekly Supply Buy List</h2>
            <p>These quantities are estimated for one person following this exact weekly plan.</p>
          </div>
          <div class="people-control" id="shopping-people-control">
            <span>Shopping people</span>
            <div class="people-stepper">
              <button type="button" id="shopping-minus">-</button>
              <strong id="shopping-people-count">1</strong>
              <button type="button" id="shopping-plus">+</button>
            </div>
          </div>
        </div>
        <div class="shopping-toggle-row">
          <button class="toggle-button" type="button" id="toggle-weekly-supply">Open Weekly Supply List</button>
        </div>
        <div class="collapsible-section" id="weekly-supply-section" hidden>
          <div class="weekly-supply-grid" id="weekly-supply-grid"></div>
        </div>
      </section>

      <section>
        <div class="section-heading">
          <div>
            <div class="section-sticker">🧺 Filtered by selected day</div>
            <h2>Shopping For This Day</h2>
            <p>Open this only when you need to buy for the currently selected day.</p>
          </div>
        </div>
        <div class="shopping-toggle-row">
          <button class="toggle-button" type="button" id="toggle-day-shopping">Open Day Shopping</button>
        </div>
        <div class="collapsible-section" id="day-shopping-section" hidden>
          <div class="weekly-supply-grid" id="daily-supply-grid"></div>
          <div class="shopping-grid" id="shopping-grid"></div>
        </div>
        <div class="shopping-toggle-row">
          <button class="toggle-button subtle" type="button" id="toggle-weekly-shopping">Open Weekly Store List</button>
        </div>
        <div class="collapsible-section" id="weekly-shopping-section" hidden>
          <div class="shopping-grid" id="weekly-shopping-grid"></div>
        </div>
      </section>

      <section class="pantry-stage">
        <div class="section-heading">
          <div>
            <h2>Pantry Basics Used In Recipes</h2>
            <p>These are the only non-workbook basics assumed anywhere on the site.</p>
          </div>
        </div>
        <div class="panel">
          <ul class="pantry-list" id="pantry-list"></ul>
        </div>
      </section>
    </section>

    <p class="footer-note">Generated from your workbook on {escape(data['generatedAt'])}. Share the public site link with anyone who needs the plan.</p>
  </main>

  <script>
    const data = {embedded_json};
    const stateKey = "diet-plan-selected-day-v1";
    const storedDay = localStorage.getItem(stateKey);
    const dishesById = Object.fromEntries(data.dishes.map((dish) => [dish.id, dish]));
    const recipesById = Object.fromEntries(data.recipes.map((recipe) => [recipe.id, recipe]));

    const dayTrack = document.getElementById("day-track");
    const mealsGrid = document.getElementById("meals-grid");
    const recipesGrid = document.getElementById("recipes-grid");
    const shoppingGrid = document.getElementById("shopping-grid");
    const dailySupplyGrid = document.getElementById("daily-supply-grid");
    const weeklySupplyGrid = document.getElementById("weekly-supply-grid");
    const weeklyShoppingGrid = document.getElementById("weekly-shopping-grid");
    const pantryList = document.getElementById("pantry-list");
    const selectedDayTitle = document.getElementById("selected-day-title");
    const selectedDayCopy = document.getElementById("selected-day-copy");
    const recipeEmpty = document.getElementById("recipe-empty");
    const dayShoppingSection = document.getElementById("day-shopping-section");
    const weeklyShoppingSection = document.getElementById("weekly-shopping-section");
    const weeklySupplySection = document.getElementById("weekly-supply-section");
    const toggleDayShopping = document.getElementById("toggle-day-shopping");
    const toggleWeeklyShopping = document.getElementById("toggle-weekly-shopping");
    const toggleWeeklySupply = document.getElementById("toggle-weekly-supply");
    const openDayShopping = document.getElementById("open-day-shopping");
    const openWeeklyShopping = document.getElementById("open-weekly-shopping");
    const recipePeopleCount = document.getElementById("recipe-people-count");
    const shoppingPeopleCount = document.getElementById("shopping-people-count");
    const recipeMinus = document.getElementById("recipe-minus");
    const recipePlus = document.getElementById("recipe-plus");
    const shoppingMinus = document.getElementById("shopping-minus");
    const shoppingPlus = document.getElementById("shopping-plus");

    let activeDayId = storedDay && data.days.some((day) => day.id === storedDay) ? storedDay : data.defaultDay;
    let activeRecipeId = null;
    let recipePeople = 1;
    let shoppingPeople = 1;

    function renderDayChips() {{
      dayTrack.innerHTML = "";
      data.days.forEach((day) => {{
        const button = document.createElement("button");
        button.className = "day-chip" + (day.id === activeDayId ? " active" : "");
        button.textContent = day.name.slice(0, 3);
        button.setAttribute("aria-label", day.name);
        button.addEventListener("click", () => {{
          activeDayId = day.id;
          localStorage.setItem(stateKey, activeDayId);
          render();
        }});
        dayTrack.appendChild(button);
      }});
    }}

    function renderMeals(day) {{
      mealsGrid.innerHTML = "";
      day.meals.forEach((meal) => {{
        const card = document.createElement("article");
        card.className = "meal-card";
        card.innerHTML = `<h4>${{meal.mealType}}</h4>`;
        const list = document.createElement("div");
        list.className = "meal-list";
        meal.dishIds.forEach((dishId) => {{
          const dish = dishesById[dishId];
          const recipe = recipesById[dish.recipeId];
        const item = document.createElement("button");
        item.className = "meal-item";
        item.type = "button";
        item.addEventListener("click", (event) => {{
          event.preventDefault();
            openRecipe(recipe.id);
        }});
          item.innerHTML = `
            <div class="meal-topline">
              <div class="dish-avatar">${{dish.avatar}}</div>
              <div>
                <strong>${{dish.icon}} ${{dish.title}}</strong>
                <span>${{recipe.summary}}</span>
              </div>
            </div>
          `;
          list.appendChild(item);
        }});
        card.appendChild(list);
        mealsGrid.appendChild(card);
      }});
    }}

    function renderRecipes(day) {{
      recipesGrid.innerHTML = "";
      const recipeIds = [...new Set(day.meals.flatMap((meal) => meal.dishIds.map((dishId) => dishesById[dishId].recipeId)))];
      if (!activeRecipeId || !recipeIds.includes(activeRecipeId)) {{
        recipeEmpty.hidden = false;
        return;
      }}
      const recipe = recipesById[activeRecipeId];
      recipeEmpty.hidden = true;
      const card = document.createElement("article");
      card.className = "panel recipe-card";
      card.id = `recipe-${{recipe.id}}`;
      const ingredients = recipe.ingredients.map((ingredient) => `
        <li>
          <strong>${{ingredient.name}}</strong>
          <span class="buy-quantity">Use: ${{ingredient.amountPlan[String(recipePeople)] || ingredient.amount}}</span>
        </li>
      `).join("");
      const steps = recipe.steps.map((step) => `<li>${{step}}</li>`).join("");
      const cookGuide = (recipe.cookGuide || []).map((guide) => `<span class="guide-pill">${{guide}}</span>`).join("");
      card.innerHTML = `
        <div class="panel-top">
          <div class="recipe-title-group">
            <div class="recipe-avatar"><span class="dish-icon">${{recipe.icon}}</span></div>
            <div>
              <h3>${{recipe.title}}</h3>
              <p>${{recipe.summary}}</p>
            </div>
          </div>
          <span class="method-pill">${{recipe.method}}</span>
        </div>
        <div class="recipe-meta">
          <span class="meta-pill">Prep ${{recipe.prepMinutes}} min</span>
          <span class="meta-pill">Cook ${{recipe.cookMinutes}} min</span>
        </div>
        ${{cookGuide ? `<div class="cook-guide">${{cookGuide}}</div>` : ""}}
        <ul class="ingredient-list">${{ingredients}}</ul>
        <ol class="steps-list">${{steps}}</ol>
        <div class="tip">${{recipe.tip}}</div>
      `;
      recipesGrid.appendChild(card);
    }}

    function renderDailySupply(day) {{
      dailySupplyGrid.innerHTML = "";
      const groups = data.dailySupply[day.id] || [];
      groups.forEach((group) => {{
        const card = document.createElement("article");
        card.className = "panel shopping-card";
        card.innerHTML = `
          <span class="store-pill">${{group.category}}</span>
          <ul class="supply-list">
            ${{group.items.map((item) => `
              <li>
                <strong>${{item.name}}</strong>
                <span class="used-quantity">Use today: ${{(item.usedPlan && item.usedPlan[String(shoppingPeople)]) || item.used}}</span>
                <span class="buy-quantity">Buy for the week: ${{(item.buyPlan && item.buyPlan[String(shoppingPeople)]) || item.buy}}</span>
                <span>Best place to buy: ${{item.store}}</span>
              </li>
            `).join("")}}
          </ul>
        `;
        dailySupplyGrid.appendChild(card);
      }});
    }}

    function renderWeeklySupply() {{
      weeklySupplyGrid.innerHTML = "";
      data.weeklySupply.forEach((group) => {{
        const card = document.createElement("article");
        card.className = "panel shopping-card";
        card.innerHTML = `
          <span class="store-pill">${{group.category}}</span>
          <ul class="supply-list">
            ${{group.items.map((item) => `
              <li>
                <strong>${{item.name}}</strong>
                <span class="buy-quantity">Buy for the week: ${{(item.buyPlan && item.buyPlan[String(shoppingPeople)]) || item.quantity}}</span>
                <span>Best place to buy: ${{item.store}}</span>
              </li>
            `).join("")}}
          </ul>
        `;
        weeklySupplyGrid.appendChild(card);
      }});
    }}

    function renderShopping(day) {{
      shoppingGrid.innerHTML = "";
      const shopping = data.dayShopping[day.id];
      const cards = [];
      shopping.stores.forEach((store) => {{
        cards.push({{ store: store.store, items: store.items }});
      }});
      if (shopping.unassigned.length) {{
        cards.push({{ store: "Needs Store Assignment", items: shopping.unassigned }});
      }}

      if (!cards.length) {{
        const empty = document.createElement("article");
        empty.className = "panel shopping-card";
        empty.innerHTML = `<h3>No shopping list mapped</h3><p>The workbook does not tie store items to this day strongly enough to filter them.</p>`;
        shoppingGrid.appendChild(empty);
        return;
      }}

      cards.forEach((store) => {{
        const card = document.createElement("article");
        card.className = "panel shopping-card";
        card.innerHTML = `
          <span class="store-pill">${{store.store}}</span>
          <ul class="shopping-list">${{store.items.map((item) => `<li>${{item}}</li>`).join("")}}</ul>
        `;
        shoppingGrid.appendChild(card);
      }});
    }}

    function renderWeeklyShopping() {{
      weeklyShoppingGrid.innerHTML = "";
      data.weeklyShopping.forEach((store) => {{
        const card = document.createElement("article");
        card.className = "panel shopping-card";
        card.innerHTML = `
          <span class="store-pill">${{store.store}}</span>
          <ul class="shopping-list">${{store.items.map((item) => `<li>${{item}}</li>`).join("")}}</ul>
        `;
        weeklyShoppingGrid.appendChild(card);
      }});
    }}

    function renderPantry() {{
      pantryList.innerHTML = data.pantryBasics.map((item) => `<li class="pantry-pill">${{item}}</li>`).join("");
    }}

    function openRecipe(recipeId) {{
      activeRecipeId = recipeId;
      render();
      const recipeNode = document.getElementById(`recipe-${{recipeId}}`);
      if (recipeNode) {{
        recipeNode.scrollIntoView({{ behavior: "smooth", block: "start" }});
        recipeNode.classList.add("flash");
        window.setTimeout(() => recipeNode.classList.remove("flash"), 1200);
      }}
    }}

    function setExpanded(section, button, expanded, openLabel, closeLabel) {{
      section.hidden = !expanded;
      button.textContent = expanded ? closeLabel : openLabel;
    }}

    function openSection(section, button, openLabel, closeLabel) {{
      setExpanded(section, button, true, openLabel, closeLabel);
      section.scrollIntoView({{ behavior: "smooth", block: "start" }});
    }}

    function closeSection(section, button, openLabel, closeLabel) {{
      setExpanded(section, button, false, openLabel, closeLabel);
    }}

    function render() {{
      const day = data.days.find((entry) => entry.id === activeDayId) || data.days[0];
      recipePeopleCount.textContent = String(recipePeople);
      shoppingPeopleCount.textContent = String(shoppingPeople);
      selectedDayTitle.textContent = day.name;
      selectedDayCopy.textContent = `${{day.name}} is focused for fast scanning: meals first, recipes next, shopping last.`;
      renderDayChips();
      renderMeals(day);
      renderRecipes(day);
      renderDailySupply(day);
      renderWeeklySupply();
      renderShopping(day);
      renderWeeklyShopping();
    }}

    toggleDayShopping.addEventListener("click", () => {{
      setExpanded(dayShoppingSection, toggleDayShopping, dayShoppingSection.hidden, "Open Day Shopping", "Close Day Shopping");
    }});
    toggleWeeklyShopping.addEventListener("click", () => {{
      setExpanded(weeklyShoppingSection, toggleWeeklyShopping, weeklyShoppingSection.hidden, "Open Weekly Store List", "Close Weekly Store List");
    }});
    toggleWeeklySupply.addEventListener("click", () => {{
      setExpanded(weeklySupplySection, toggleWeeklySupply, weeklySupplySection.hidden, "Open Weekly Supply List", "Close Weekly Supply List");
    }});
    openDayShopping.addEventListener("click", () => {{
      closeSection(weeklySupplySection, toggleWeeklySupply, "Open Weekly Supply List", "Close Weekly Supply List");
      closeSection(weeklyShoppingSection, toggleWeeklyShopping, "Open Weekly Store List", "Close Weekly Store List");
      openSection(dayShoppingSection, toggleDayShopping, "Open Day Shopping", "Close Day Shopping");
    }});
    openWeeklyShopping.addEventListener("click", () => {{
      closeSection(dayShoppingSection, toggleDayShopping, "Open Day Shopping", "Close Day Shopping");
      closeSection(weeklyShoppingSection, toggleWeeklyShopping, "Open Weekly Store List", "Close Weekly Store List");
      openSection(weeklySupplySection, toggleWeeklySupply, "Open Weekly Supply List", "Close Weekly Supply List");
    }});
    recipeMinus.addEventListener("click", () => {{
      recipePeople = Math.max(1, recipePeople - 1);
      render();
    }});
    recipePlus.addEventListener("click", () => {{
      recipePeople = Math.min(6, recipePeople + 1);
      render();
    }});
    shoppingMinus.addEventListener("click", () => {{
      shoppingPeople = Math.max(1, shoppingPeople - 1);
      render();
    }});
    shoppingPlus.addEventListener("click", () => {{
      shoppingPeople = Math.min(6, shoppingPeople + 1);
      render();
    }});

    renderPantry();
    render();
  </script>
</body>
</html>
"""


def main():
    data = build_data()
    DATA_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
    SITE_PATH.write_text(build_html(data), encoding="utf-8")
    print(f"Wrote {SITE_PATH}")
    print(f"Wrote {DATA_PATH}")


if __name__ == "__main__":
    main()
